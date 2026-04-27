"""엑셀 I/O 관리 (openpyxl 기반).

관대한 로드 (tolerant load) 정책:
    - 스키마가 어긋나거나 행 데이터가 검증을 통과하지 못해도 예외로 튕기지 않는다.
    - 모든 행을 읽어서 UI에 그대로 보여준다.
    - 유효한 행은 Order 로, 유효하지 않은 행은 RawRow (에러 메시지 포함) 로 반환한다.
    - 사용자가 UI에서 값을 수정하면 재검증되어 Order 로 승격될 수 있다.

입력 스키마 (8컬럼):
    구매처 / 수취인 / 수취인번호 / 통관번호 / 우편번호 / 수취인 주소 / 수량 / 영문이름

출력 스키마 (10컬럼):
    입력 8컬럼 + 토탈가격 + 주문번호
"""
from __future__ import annotations

import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import ValidationError

from src.exceptions import ExcelError
from src.models.order import Order
from src.utils.logger import get_logger

log = get_logger()

# 입력 엑셀 필수 컬럼 (스크린샷 순서와 동일)
# "수취인번호.1" 은 보조 번호 — 자동화에는 사용 안 함, 단순 보존.
REQUIRED_COLUMNS = (
    "구매처",
    "수취인",
    "수취인번호",
    "수취인번호.1",
    "통관번호",
    "우편번호",
    "수취인 주소",
    "수량",
    "영문이름",
)

# 출력 엑셀 추가 컬럼
OUTPUT_EXTRA_COLUMNS = ("토탈가격", "주문번호")

ALL_COLUMNS = REQUIRED_COLUMNS + OUTPUT_EXTRA_COLUMNS

# 입력 컬럼 → Order 필드 매핑
_INPUT_FIELD_MAP = {
    "구매처": "product_url",
    "수취인": "name",
    "수취인번호": "phone",
    "수취인번호.1": "phone2",
    "통관번호": "customs_id",
    "우편번호": "postal_code",
    "수취인 주소": "address",
    "수량": "quantity",
    "영문이름": "english_name",
}

# 과거 스키마 별칭 (하위 호환)
_LEGACY_ALIASES = {
    "구매처": ("상품링크",),
    "수취인": ("이름", "받는사람"),
    "수취인번호": ("번호", "연락처"),
    "수취인 주소": ("주소",),
}

# 상태별 색상
_STATUS_FILL = {
    "completed": PatternFill("solid", fgColor="C6EFCE"),
    "in_progress": PatternFill("solid", fgColor="BDD7EE"),
    "paused": PatternFill("solid", fgColor="FFF2CC"),
    "failed": PatternFill("solid", fgColor="FFC7CE"),
    "pending": PatternFill("solid", fgColor="F2F2F2"),
    "invalid": PatternFill("solid", fgColor="FDE2E2"),
    "unavailable": PatternFill("solid", fgColor="E5E7EB"),  # 회색 — 판매 불가
}


@dataclass
class RawRow:
    """검증을 통과하지 못한 엑셀 행.

    UI 에는 표 형태로 그대로 보여주되, 편집 시 다시 검증을 시도해 통과하면
    Order 로 승격한다. 사용자가 값을 수정 → 저장하면 원본 엑셀에 반영된다.
    """

    row: int
    fields: dict[str, str] = field(default_factory=dict)  # 입력 컬럼명 → 문자열 값
    total_price: int | None = None
    order_number: str | None = None
    error: str = ""

    # UI 에서 Order 와 동일하게 다루기 위한 속성들
    status: str = "invalid"

    def get(self, col_name: str) -> str:
        return self.fields.get(col_name, "")

    def set(self, col_name: str, value: str) -> None:
        self.fields[col_name] = value

    def is_done(self) -> bool:
        return False

    def is_retryable(self) -> bool:
        return False

    def needs_price(self) -> bool:
        return self.total_price is None


# 테이블/저장 등에서 섞어서 다루는 타입
RowItem = Order | RawRow


class ExcelManager:
    """엑셀 파일 1개에 대한 read/write 래퍼."""

    def __init__(self, path: str | Path):
        self.path = Path(path)
        self._rows: list[RowItem] = []
        self._output_path: Path | None = None
        # 실제 파일에서 읽은 헤더 순서 (원본 덮어쓰기 시 유지)
        self._header_order: list[str] = []

    # -------------------------------------------------------------
    # Load (tolerant)
    # -------------------------------------------------------------

    def load(self, backup: bool = True) -> list[RowItem]:
        """엑셀을 관대하게 로드. 예외는 파일 자체를 열 수 없을 때만 던진다."""
        if not self.path.exists():
            raise ExcelError(f"엑셀 파일을 찾을 수 없습니다: {self.path}")

        if backup:
            self._backup_original()

        try:
            wb = load_workbook(self.path, data_only=True)
        except Exception as exc:
            raise ExcelError(f"엑셀 파일 열기 실패: {exc}") from exc

        ws = wb.active
        if ws is None:
            wb.close()
            self._rows = []
            return self._rows

        # 헤더 읽기 + 별칭 해석
        header_row = [str(c.value).strip() if c.value else "" for c in ws[1]]
        self._header_order = [h for h in header_row if h]

        col_idx: dict[str, int] = {}
        for i, name in enumerate(header_row):
            if not name:
                continue
            col_idx[name] = i
            for canonical, aliases in _LEGACY_ALIASES.items():
                if name in aliases and canonical not in col_idx:
                    col_idx[canonical] = i

        missing_headers = [c for c in REQUIRED_COLUMNS if c not in col_idx]
        if missing_headers:
            log.warning(
                f"엑셀 헤더에 누락된 필수 컬럼이 있습니다 (UI에서 수정 가능): {missing_headers}"
            )

        rows: list[RowItem] = []
        if ws.max_row is None or ws.max_row < 2:
            wb.close()
            self._rows = rows
            log.info(f"엑셀 로드 완료: 0건 (헤더만 또는 빈 시트)")
            return rows

        for row_idx, raw_row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 완전히 빈 행 스킵
            if all(
                v is None or (isinstance(v, str) and not v.strip()) for v in raw_row
            ):
                continue

            fields_str = self._extract_fields(raw_row, col_idx)
            total_price = self._extract_total_price(raw_row, col_idx)
            order_number = self._extract_order_number(raw_row, col_idx)

            item = self._try_build_order(
                row_idx, fields_str, total_price, order_number
            )
            if item is None:
                # 검증 실패 → RawRow 로 담아서라도 UI 에 노출
                err_msg = self._last_error or "검증 실패"
                item = RawRow(
                    row=row_idx,
                    fields=fields_str,
                    total_price=total_price,
                    order_number=order_number,
                    error=err_msg,
                )
            rows.append(item)

        wb.close()
        self._rows = rows

        valid = sum(1 for r in rows if isinstance(r, Order))
        invalid = len(rows) - valid
        log.info(
            f"엑셀 로드 완료: 총 {len(rows)}건 (유효 {valid} / 수정필요 {invalid}) — {self.path}"
        )
        return rows

    _last_error: str = ""

    def _extract_fields(self, row: tuple, col_idx: dict[str, int]) -> dict[str, str]:
        """입력 8컬럼 값을 문자열 딕셔너리로 추출. 누락 컬럼은 빈 문자열.

        엑셀 셀이 숫자형일 때 openpyxl은 int/float를 넘겨준다.
        특히 우편번호/수량은 숫자형으로 저장되기 쉬우므로 '6236.0' 같은
        지저분한 변환을 피하기 위해 정수로 표현 가능한 float는 int로 변환한다.
        """
        out: dict[str, str] = {}
        for col_name in REQUIRED_COLUMNS:
            i = col_idx.get(col_name)
            if i is None or i >= len(row):
                out[col_name] = ""
                continue
            v = row[i]
            if v is None:
                out[col_name] = ""
            elif isinstance(v, float) and v.is_integer():
                out[col_name] = str(int(v))
            else:
                out[col_name] = str(v).strip()
        return out

    def _extract_total_price(
        self, row: tuple, col_idx: dict[str, int]
    ) -> int | None:
        i = col_idx.get("토탈가격")
        if i is None or i >= len(row):
            return None
        v = row[i]
        if v is None or v == "":
            return None
        try:
            return int(str(v).replace(",", "").replace("원", "").strip())
        except ValueError:
            return None

    def _extract_order_number(
        self, row: tuple, col_idx: dict[str, int]
    ) -> str | None:
        i = col_idx.get("주문번호")
        if i is None or i >= len(row):
            return None
        v = row[i]
        if v is None or str(v).strip() == "":
            return None
        return str(v).strip()

    def _try_build_order(
        self,
        row_idx: int,
        fields: dict[str, str],
        total_price: int | None,
        order_number: str | None,
    ) -> Order | None:
        data: dict = {"row": row_idx}
        for col_name, field_name in _INPUT_FIELD_MAP.items():
            raw = fields.get(col_name, "")
            if field_name == "quantity":
                data[field_name] = raw
            else:
                data[field_name] = raw

        if total_price is not None:
            data["total_price"] = total_price
            qty_raw = fields.get("수량", "1")
            try:
                qty_int = int(float(qty_raw)) if qty_raw else 1
                if qty_int >= 1 and total_price % qty_int == 0:
                    data["unit_price"] = total_price // qty_int
            except (TypeError, ValueError):
                pass

        if order_number:
            data["order_number"] = order_number
            data["status"] = "completed"

        try:
            return Order.model_validate(data)
        except ValidationError as exc:
            self._last_error = self._format_validation_error(exc)
            return None
        except Exception as exc:
            self._last_error = str(exc)
            return None

    @staticmethod
    def _format_validation_error(exc: ValidationError) -> str:
        parts = []
        for err in exc.errors():
            loc = ".".join(str(x) for x in err["loc"])
            parts.append(f"{loc}: {err['msg']}")
        return "; ".join(parts)

    def _backup_original(self) -> Path:
        backup_dir = Path("data/backups")
        backup_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"{self.path.stem}_BACKUP_{ts}{self.path.suffix}"
        shutil.copy2(self.path, backup_path)
        log.info(f"원본 백업: {backup_path}")
        return backup_path

    # -------------------------------------------------------------
    # Save
    # -------------------------------------------------------------

    @property
    def output_path(self) -> Path:
        if self._output_path is None:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            self._output_path = self.path.with_name(
                f"{self.path.stem}_완료_{ts}{self.path.suffix}"
            )
        return self._output_path

    def save(
        self,
        rows: list[RowItem] | None = None,
        target: Path | None = None,
    ) -> Path:
        """결과 파일 저장 (새 타임스탬프 파일로)."""
        rows = rows if rows is not None else self._rows
        out = Path(target) if target else self.output_path
        return self._write_workbook(rows, out)

    def save_to_original(self, rows: list[RowItem] | None = None) -> Path:
        """원본 엑셀 파일에 덮어쓰기 저장.

        UI에서 수정한 내용을 원본에 반영할 때 사용. 저장 직전에 현재 원본을
        data/backups/ 로 자동 백업한다.
        """
        rows = rows if rows is not None else self._rows
        if self.path.exists():
            self._backup_original()
        return self._write_workbook(rows, self.path)

    def _write_workbook(self, rows: list[RowItem], out: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "주문 목록"

        # 헤더
        for i, col_name in enumerate(ALL_COLUMNS, start=1):
            c = ws.cell(row=1, column=i, value=col_name)
            c.font = Font(bold=True)

        for i, item in enumerate(rows, start=2):
            values = self._row_values(item)
            for col_i, val in enumerate(values, start=1):
                ws.cell(row=i, column=col_i, value=val)

            status = getattr(item, "status", "pending")
            fill = _STATUS_FILL.get(status)
            if fill:
                for col in range(1, len(ALL_COLUMNS) + 1):
                    ws.cell(row=i, column=col).fill = fill

        widths = {
            1: 50, 2: 12, 3: 16, 4: 16, 5: 10,
            6: 40, 7: 8, 8: 20, 9: 14, 10: 18,
        }
        for idx, w in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = w

        tmp = out.with_suffix(out.suffix + ".tmp")
        try:
            wb.save(tmp)
            tmp.replace(out)
        except Exception as exc:
            raise ExcelError(f"엑셀 저장 실패: {exc}") from exc
        finally:
            wb.close()

        log.info(f"엑셀 저장: {out}")
        return out

    def _row_values(self, item: RowItem) -> list:
        """Order 또는 RawRow 를 출력 컬럼 순서대로 값 리스트로 변환."""
        if isinstance(item, Order):
            return [
                item.product_url,
                item.name,
                item.phone,
                item.customs_id,
                item.postal_code,
                item.address,
                item.quantity,
                item.english_name,
                item.total_price if item.total_price is not None else "",
                item.order_number or "",
            ]
        # RawRow
        return [
            item.get("구매처"),
            item.get("수취인"),
            item.get("수취인번호"),
            item.get("통관번호"),
            item.get("우편번호"),
            item.get("수취인 주소"),
            item.get("수량"),
            item.get("영문이름"),
            item.total_price if item.total_price is not None else "",
            item.order_number or "",
        ]

    # -------------------------------------------------------------
    # Helpers for UI
    # -------------------------------------------------------------

    def update_order(self, order: Order, autosave: bool = False) -> None:
        """메모리 상의 주문 갱신.

        autosave=True 일 때만 디스크에 저장한다. 기본값은 False.
        매 행마다 디스크 I/O 가 일어나면 qasync 이벤트루프를 점유해
        Playwright IPC 와 충돌(reentry RuntimeError)이 생기므로,
        호출 측에서 적절한 시점(가격조회 끝/주문 완료/주기적)에만 저장한다.
        """
        for i, r in enumerate(self._rows):
            if getattr(r, "row", None) == order.row:
                self._rows[i] = order
                break
        else:
            self._rows.append(order)
        if autosave:
            self.save(self._rows)

    def replace_rows(self, rows: list[RowItem]) -> None:
        """UI에서 편집된 리스트로 교체 (저장은 안 함)."""
        self._rows = list(rows)

    @property
    def rows(self) -> list[RowItem]:
        return list(self._rows)

    @property
    def orders(self) -> list[Order]:
        """유효한 Order만 반환 (자동화에 쓰이는 지점용)."""
        return [r for r in self._rows if isinstance(r, Order)]

    def try_promote(self, row: int, fields: dict[str, str]) -> RowItem:
        """편집된 값으로 RawRow 를 Order 로 승격 시도. 실패하면 RawRow 반환."""
        tp: int | None = None
        existing = next((r for r in self._rows if getattr(r, "row", None) == row), None)
        order_no = None
        if existing is not None:
            tp = getattr(existing, "total_price", None)
            order_no = getattr(existing, "order_number", None)
        order = self._try_build_order(row, fields, tp, order_no)
        if order is not None:
            return order
        return RawRow(
            row=row,
            fields=fields,
            total_price=tp,
            order_number=order_no,
            error=self._last_error or "검증 실패",
        )
