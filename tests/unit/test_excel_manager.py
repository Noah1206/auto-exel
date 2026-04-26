"""ExcelManager 단위 테스트 (tolerant load + save-back)."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from src.core.excel_manager import (
    ALL_COLUMNS,
    ExcelManager,
    RawRow,
    REQUIRED_COLUMNS,
)
from src.exceptions import ExcelError
from src.models.order import Order


def _build_sample_excel(path: Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    headers = list(REQUIRED_COLUMNS)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, h in enumerate(headers, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(h, ""))
    wb.save(path)
    wb.close()


def _sample_rows() -> list[dict]:
    return [
        {
            "구매처": "https://www.11st.co.kr/products/1",
            "수취인": "김철수",
            "수취인번호": "010-1234-5678",
            "통관번호": "P123456789012",
            "우편번호": "06236",
            "수취인 주소": "서울시 강남구",
            "수량": 2,
            "영문이름": "KIM CHUL SOO",
        },
        {
            "구매처": "https://www.11st.co.kr/products/2",
            "수취인": "이영희",
            "수취인번호": "01023456789",
            "통관번호": "P987654321098",
            "우편번호": "48094",
            "수취인 주소": "부산시 해운대구",
            "수량": 1,
            "영문이름": "LEE YOUNG HEE",
        },
    ]


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "sample.xlsx"
    _build_sample_excel(path, _sample_rows())
    return path


def test_load_valid(sample_xlsx, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    mgr = ExcelManager(sample_xlsx)
    rows = mgr.load(backup=False)
    assert len(rows) == 2
    assert all(isinstance(r, Order) for r in rows)
    assert rows[0].name == "김철수"
    assert rows[0].quantity == 2
    assert rows[1].phone == "010-2345-6789"


def test_tolerant_load_missing_columns(tmp_path, monkeypatch):
    """필수 컬럼이 누락되어도 예외 없이 행을 RawRow로 반환한다."""
    monkeypatch.chdir(tmp_path)
    path = tmp_path / "bad.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["구매처", "수취인"])  # 나머지 누락
    ws.append(["https://www.11st.co.kr/x", "김"])
    wb.save(path)
    wb.close()
    rows = ExcelManager(path).load(backup=False)
    assert len(rows) == 1
    assert isinstance(rows[0], RawRow)
    assert rows[0].get("구매처") == "https://www.11st.co.kr/x"
    assert rows[0].get("수취인") == "김"
    assert rows[0].error  # 에러 메시지 포함


def test_tolerant_load_invalid_row(tmp_path, monkeypatch):
    """행 값 검증 실패도 RawRow로 반환되어 UI에 노출된다."""
    monkeypatch.chdir(tmp_path)
    path = tmp_path / "mixed.xlsx"
    _build_sample_excel(
        path,
        [
            {
                "구매처": "https://www.11st.co.kr/products/1",
                "수취인": "김철수",
                "수취인번호": "010-1234-5678",
                "통관번호": "BADVAL",  # 잘못된 통관번호
                "우편번호": "06236",
                "수취인 주소": "서울",
                "수량": 1,
                "영문이름": "KIM",
            },
            {  # 정상 행
                "구매처": "https://www.11st.co.kr/products/2",
                "수취인": "이",
                "수취인번호": "01023456789",
                "통관번호": "P987654321098",
                "우편번호": "48094",
                "수취인 주소": "부산",
                "수량": 1,
                "영문이름": "LEE",
            },
        ],
    )
    rows = ExcelManager(path).load(backup=False)
    assert len(rows) == 2
    assert isinstance(rows[0], RawRow)
    assert isinstance(rows[1], Order)


def test_load_file_not_found(tmp_path):
    mgr = ExcelManager(tmp_path / "nope.xlsx")
    with pytest.raises(ExcelError):
        mgr.load(backup=False)


def test_try_promote_from_raw(tmp_path, monkeypatch):
    """RawRow 상태에서 사용자가 값을 고치면 Order로 승격된다."""
    monkeypatch.chdir(tmp_path)
    path = tmp_path / "mixed.xlsx"
    _build_sample_excel(
        path,
        [
            {
                "구매처": "https://www.11st.co.kr/products/1",
                "수취인": "김철수",
                "수취인번호": "010-1234-5678",
                "통관번호": "BADVAL",
                "우편번호": "06236",
                "수취인 주소": "서울",
                "수량": 1,
                "영문이름": "KIM",
            }
        ],
    )
    mgr = ExcelManager(path)
    rows = mgr.load(backup=False)
    assert isinstance(rows[0], RawRow)

    fixed = mgr.try_promote(
        rows[0].row,
        {
            "구매처": "https://www.11st.co.kr/products/1",
            "수취인": "김철수",
            "수취인번호": "010-1234-5678",
            "통관번호": "P123456789012",  # 수정됨
            "우편번호": "06236",
            "수취인 주소": "서울",
            "수량": "1",
            "영문이름": "KIM",
        },
    )
    assert isinstance(fixed, Order)
    assert fixed.customs_id == "P123456789012"


def test_excel_numeric_postal_recovered(tmp_path, monkeypatch):
    """엑셀에서 06236 이 숫자로 저장돼 6236으로 들어와도 복원된다."""
    monkeypatch.chdir(tmp_path)
    path = tmp_path / "numeric_postal.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(list(REQUIRED_COLUMNS))
    ws.append(
        [
            "https://www.11st.co.kr/products/1",
            "김철수",
            "010-1234-5678",
            "P123456789012",
            6236,  # ← 숫자형, 앞자리 0 누락
            "서울",
            1,
            "KIM",
        ]
    )
    wb.save(path)
    wb.close()
    rows = ExcelManager(path).load(backup=False)
    assert isinstance(rows[0], Order)
    assert rows[0].postal_code == "06236"


def test_excel_old_6digit_postal_rejected(tmp_path, monkeypatch):
    """구 6자리 우편번호는 RawRow로 남고 에러 메시지에 안내가 포함된다."""
    monkeypatch.chdir(tmp_path)
    path = tmp_path / "old_postal.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(list(REQUIRED_COLUMNS))
    ws.append(
        [
            "https://www.11st.co.kr/products/1",
            "김철수",
            "010-1234-5678",
            "P123456789012",
            "123-456",  # 구 6자리
            "서울",
            1,
            "KIM",
        ]
    )
    wb.save(path)
    wb.close()
    rows = ExcelManager(path).load(backup=False)
    assert isinstance(rows[0], RawRow)
    assert "6자리" in rows[0].error or "폐지" in rows[0].error


def test_legacy_column_aliases(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    path = tmp_path / "legacy.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(
        ["상품링크", "이름", "번호", "통관번호", "우편번호", "주소", "수량", "영문이름"]
    )
    ws.append(
        [
            "https://www.11st.co.kr/x",
            "김철수",
            "010-1234-5678",
            "P123456789012",
            "06236",
            "서울시 강남구",
            1,
            "KIM CHUL SOO",
        ]
    )
    wb.save(path)
    wb.close()
    rows = ExcelManager(path).load(backup=False)
    assert len(rows) == 1
    assert isinstance(rows[0], Order)
    assert rows[0].name == "김철수"


def test_save_stage1_price_only(sample_xlsx, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    mgr = ExcelManager(sample_xlsx)
    rows = mgr.load(backup=False)
    rows[0].unit_price = 10000
    rows[0].compute_total()
    out = mgr.save(rows)
    assert out.exists()

    wb = load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert list(header) == list(ALL_COLUMNS)
    total_col = header.index("토탈가격") + 1
    order_no_col = header.index("주문번호") + 1
    assert ws.cell(row=2, column=total_col).value == 20000
    assert ws.cell(row=2, column=order_no_col).value in (None, "")
    wb.close()


def test_save_stage2_with_order_number(sample_xlsx, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    mgr = ExcelManager(sample_xlsx)
    rows = mgr.load(backup=False)
    rows[0].unit_price = 10000
    rows[0].compute_total()
    rows[0].status = "completed"
    rows[0].order_number = "202604220001"
    out = mgr.save(rows)

    wb = load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    total_col = header.index("토탈가격") + 1
    order_no_col = header.index("주문번호") + 1
    assert ws.cell(row=2, column=total_col).value == 20000
    assert ws.cell(row=2, column=order_no_col).value == "202604220001"
    wb.close()


def test_save_to_original(sample_xlsx, tmp_path, monkeypatch):
    """원본 파일에 덮어쓰기 저장이 동작한다."""
    monkeypatch.chdir(tmp_path)
    mgr = ExcelManager(sample_xlsx)
    rows = mgr.load(backup=False)
    # UI에서 수정한 효과: 수취인 이름 변경
    rows[0].name = "김수정"
    out = mgr.save_to_original(rows)
    assert out == sample_xlsx

    mgr2 = ExcelManager(sample_xlsx)
    reloaded = mgr2.load(backup=False)
    assert isinstance(reloaded[0], Order)
    assert reloaded[0].name == "김수정"


def test_save_to_original_writes_raw_rows(tmp_path, monkeypatch):
    """RawRow 상태의 수정사항도 원본에 그대로 기록된다."""
    monkeypatch.chdir(tmp_path)
    path = tmp_path / "mixed.xlsx"
    _build_sample_excel(
        path,
        [
            {
                "구매처": "https://www.11st.co.kr/products/1",
                "수취인": "김철수",
                "수취인번호": "010-1234-5678",
                "통관번호": "BADVAL",  # invalid
                "우편번호": "06236",
                "수취인 주소": "서울",
                "수량": 1,
                "영문이름": "KIM",
            }
        ],
    )
    mgr = ExcelManager(path)
    rows = mgr.load(backup=False)
    assert isinstance(rows[0], RawRow)
    # 사용자가 UI에서 통관번호를 바꿨지만 아직 다른 필드 이슈로 RawRow 유지된 상태라 가정
    rows[0].set("통관번호", "P111222333444")
    rows[0].set("수량", "")  # 수량은 비워서 아직 raw 유지 가정
    mgr.save_to_original(rows)

    # 원본을 다시 읽어서 통관번호 수정이 반영됐는지 확인
    mgr2 = ExcelManager(path)
    reloaded = mgr2.load(backup=False)
    # 수량 빈 값이라 아직 RawRow 가능성이 높음 → 수정된 통관번호는 남아있어야 함
    r = reloaded[0]
    if isinstance(r, RawRow):
        assert r.get("통관번호") == "P111222333444"
    else:
        assert r.customs_id == "P111222333444"
